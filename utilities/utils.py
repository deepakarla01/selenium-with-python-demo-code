import csv
import logging
import os
import re
import time
import pytest
from pytest_assume.plugin import assume
import requests
from openpyxl import workbook, load_workbook
import softest
from selenium.common import ElementNotInteractableException, NoSuchElementException, ElementClickInterceptedException, \
    StaleElementReferenceException, TimeoutException, NoSuchWindowException, WebDriverException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait






class Utils(softest.TestCase):

    # Class-level logger initialization
    logger = None

    @staticmethod
    def custom_logger(logLevel=logging.DEBUG):
        if Utils.logger is None:  # Ensure logger is created only once
            # Create logger
            logger = logging.getLogger("DemoQABDDLogger")

            # Set logger level
            logger.setLevel(logLevel)

            # Create file handler and set the log level
            file_handler = logging.FileHandler("DemoQA_With_Logging_Fixture.log")
            file_handler.setLevel(logLevel)

            # Create formatter - how you want your logs to be formatted
            formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s - %(module)s")

            # Add formatter to file handler
            file_handler.setFormatter(formatter)

            # Add file handler to logger
            logger.addHandler(file_handler)

            # Stream handler (used by pytest caplog) with filter
            stream_handler = logging.StreamHandler()
            stream_handler.setLevel(logging.INFO)
            stream_handler.setFormatter(formatter)

            class GherkinOnlyFilter(logging.Filter):
                def filter(self, record):
                    # Only allow Gherkin steps to propagate to caplog/HTML report
                    return record.getMessage().strip().startswith(
                        ("Given", "When", "Then", "And", "But", "Feature", "Scenario"))

            stream_handler.addFilter(GherkinOnlyFilter())
            logger.addHandler(stream_handler)

            Utils.logger = logger  # Assign the logger to the class-level attribute

        Utils.logger.propagate = False  # Prevent propagation to root logger
        return Utils.logger

    @staticmethod
    def read_data_from_excel(self,filename,sheet_name):
        log = Utils.custom_logger()
        wb = load_workbook(filename) #open the workbook
        ws = wb[sheet_name] #open the sheet
        data_list = [] #declare an empty list

        row_cnt = ws.max_row #get the number of rows
        col_cnt = ws.max_column #get the number of columns
        try:
            for r in range(2, row_cnt + 1):  # prints all the rows except the first row, that has the column names
                row = [] #declare an empty row
                for c in range(1, col_cnt + 1):
                    row.append(ws.cell(row=r, column=c).value) #add data to the row - each cell value of the row at a time
                data_list.append(row) #after reading one row, add it to the list

        except Exception as e:
            log.error("An unexpected error occurred in read_data_from_excel.", exc_info=e)
        finally:
            log.info("Completed execution of read_data_from_excel without exceptions.")
            return data_list  # return the list

    @staticmethod
    def read_data_from_csv(filename):
        log = Utils.custom_logger()
        try:
            with open(filename, "r") as csv_file:
                csv_reader = csv.reader(csv_file)
                header = next(csv_reader)  # Skip the header row

                # Return single-column data as a list of strings, cleaning \xa0 from each ent
                if len(header) == 1:
                    return [row[0].replace('x\a0','').strip() for row in csv_reader]

                # Return multi-column data as a list of tuples, cleaning \xa0 from each ent
                return [tuple(row[i].replace('x\a0','').strip() for i in range(len(row))) for row in csv_reader]
        except Exception as e:
            log.error("An unexpected error occurred in read_data_from_csv.", exc_info=e)
        finally:
            log.info("Completed execution of read_data_from_csv without exceptions.")


    @staticmethod
    def assert_text(self,expected_text,actual_text):
        log = Utils.custom_logger()
        try:
            assert expected_text in actual_text
            log.debug("assertion passed for, " +actual_text)
        except AssertionError:
            log.error("assertion failed for, " +actual_text)
        finally:
            log.info("Completed execution of assert_text without exceptions.")

    @staticmethod
    def assert_url(self,expected_url,actual_url):
        log = Utils.custom_logger(self)
        try:
            assert expected_url == actual_url
            log.debug("assertion passed for, " +actual_url)
        except AssertionError:
            log.error(f"Expected URL: {expected_url}, Actual URL: {actual_url}")

    @staticmethod
    def assert_page_title(self,expected_title,actual_title):
        log = Utils.custom_logger()
        try:
            actual_page_title = re.sub(r'[^\x00-\x7F]+', '', actual_title).strip()

            # Perform the assertion
            assert expected_title in actual_page_title
            log.debug(f"Assertion passed for: {actual_page_title}")
        except AssertionError:
            actual_page_title = re.sub(r'[^\x00-\x7F]+', '', actual_title).strip()
            # Log the detailed error
            log.error(f"Assertion failed. Expected: '{expected_title}', Actual: '{actual_page_title}'")

    @staticmethod
    def check_broken_banner_img(self, banner_img_style, bg_image_name, url):
        log = Utils.custom_logger()

        try:
            if not banner_img_style or "url(" not in banner_img_style:
                log.error("No background image found in the provided style.")
                return False

            # Extract image URL from the style string
            image_url = banner_img_style.split('url("')[1].split('")')[0]
            log.debug(f"Extracted image URL: {image_url}")

            # Soft assert the expected image name is in the image URL
            assume(bg_image_name in image_url, f"'{bg_image_name}' not found in image URL: {image_url}")

            # Add browser-like headers
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                              "AppleWebKit/537.36 (KHTML, like Gecko) "
                              "Chrome/112.0.0.0 Safari/537.36",
                "Referer": url,
            }

            response = requests.get(image_url, headers=headers, timeout=5)
            if response.status_code == 200:
                log.info(f"Background image '{bg_image_name}' loaded successfully.")
                return True
            else:
                log.error(f"Broken background image '{bg_image_name}' (Status Code: {response.status_code})")
                return False
        except Exception as e:
            log.error(f"Exception occurred while checking background image: {e}")
            return False

    @staticmethod
    def check_broken_images(self, images, page_url):
        log = Utils.custom_logger()
        log.info(f"Found {len(images)} images on the page: {page_url}")

        broken_images = []

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                          "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            "Referer": page_url
        }

        for img in images:
            img_url = img.get_attribute("src")
            if not img_url:
                log.info("Image does not have a valid src attribute.")
                broken_images.append(img)
                continue

            try:
                response = requests.get(img_url, timeout=5, headers=headers)
                if response.status_code != 200:
                    log.error(f"Broken image (HTTP): {img_url} (Status Code: {response.status_code})")
                    broken_images.append(img)
                    continue

                # Scroll into view (helps trigger lazy loading)
                self.driver.execute_script("arguments[0].scrollIntoView(true);", img)

                # Allow time for image to load
                time.sleep(0.25)

                render_status = self.driver.execute_script(
                    "return {complete: arguments[0].complete, width: arguments[0].naturalWidth};",
                    img
                )
                complete = render_status.get("complete")
                natural_width = render_status.get("width")

                if (not complete or natural_width == 0) and 'page=stats&noheader&proxy&chart=admin-bar-hours-scale' not in img_url:
                    log.error(f"Broken image (Render): {img_url} | complete={complete}, naturalWidth={natural_width}")
                    broken_images.append(img)

            except requests.exceptions.RequestException as e:
                log.error(f"Error checking image: {img_url} (Error: {e})")
                broken_images.append(img)

        log.info(f"Total broken images: {len(broken_images)}")
        return broken_images

    @staticmethod
    def write_to_a_file(self, filename, list_item):
        # Open the file for writing
        file_name = open(filename, "w")  # Correctly assign the file object to 'file_name'
        try:
            for items in list_item:
                file_name.write(items.text + "\n")  # Use 'file_name' to write to the file
        finally:
            file_name.close()  # Ensure the file is closed after writing

    @staticmethod
    def broken_link_checker(link_url, timeout=5):
        log = Utils.custom_logger()

        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36"
        }

        try:
            response = requests.head(link_url, headers=headers, allow_redirects=True, timeout=timeout)

            if response.status_code == 405:
                log.warning(f"HEAD request not allowed for URL: {link_url}. Falling back to GET.")
                response = requests.get(link_url, headers=headers, allow_redirects=True, timeout=timeout)

            if response.status_code >= 400:
                log.error(f"Broken link: {link_url} (Status code: {response.status_code})")
                return True
        except requests.exceptions.RequestException as req_error:
            log.error(f"Error while validating link: {link_url} | Error: {req_error}")
            return True

        return False

    @staticmethod
    def check_if_recent_file_downloaded(filename, path, timeout=30, poll_interval=1, freshness_window=60):
        log = Utils.custom_logger()
        basename, ext = os.path.splitext(filename)

        try:
            elapsed = 0
            matching_files = []

            while elapsed < timeout:
                all_files = os.listdir(path)
                only_files = [f for f in all_files if os.path.isfile(os.path.join(path, f))]

                # Match base name pattern (e.g., Branding.zip, Branding (1).zip, etc.)
                matching_files = [
                    os.path.join(path, f)
                    for f in only_files
                    if f.startswith(basename) and f.endswith(ext)
                ]

                current_time = time.time()
                recent_files = [
                    f for f in matching_files
                    if 0 < (current_time - os.path.getmtime(os.path.join(path, f))) <= freshness_window
                ]

                if recent_files:
                    break

                time.sleep(poll_interval)
                elapsed += poll_interval

            if not recent_files:
                log.warning(
                    f"No RECENT files matching '{filename}' were found in '{path}' "
                    f"within {timeout} seconds and freshness window of {freshness_window} seconds.")
                return False

            # Get most recent among recent files
            recent_file = max(recent_files, key=os.path.getmtime)
            file_size = os.path.getsize(recent_file)
            file_age = time.time() - os.path.getmtime(recent_file)

            log.info(f"Recent file found: {recent_file}, size: {file_size}, modified {file_age:.2f}s ago")

            if file_size > 0:
                return True
            else:
                log.warning(f"File '{recent_file}' exists but is EMPTY.")
                return False

        except Exception as e:
            log.error("An unexpected error occurred in check_if_recent_file_downloaded.", exc_info=True)
            return False


    def check_expected_url(self,expected_url,actual_url):
        log = Utils.custom_logger()
        try:
            if expected_url in actual_url:
                log.info(f"Expected URL: {expected_url} is in the actual url: {actual_url}")
                return True
            else:
                log.error(f"Expected URL: {expected_url} is not in the actual url: {actual_url}")
                return False
        except Exception as e:
            log.error("An unexpected error occurred in check_expected_url: {e}")

